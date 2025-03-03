#!/usr/bin/env python3

"""
===========================
= AWS RESOURCE SCANNER =
===========================

Title: AWS Account Billing Data Export
Version: v1.1.0
Date: MAR-03-2025

Description:
Exports AWS billing data for specified time periods (monthly or yearly), 
organized by service and associated cost. Handles AWS Cost Explorer
limitations for historical data access.
"""

import os
import sys
import datetime
import re
import boto3
from dateutil.relativedelta import relativedelta
from botocore.exceptions import ClientError
from pathlib import Path

# Add path to import utils module
try:
    # Try to import directly (if utils.py is in Python path)
    import utils
except ImportError:
    # If import fails, try to find the module relative to this script
    script_dir = Path(__file__).parent.absolute()
    
    # Check if we're in the scripts directory
    if script_dir.name.lower() == 'scripts':
        # Add the parent directory (StratusScan root) to the path
        sys.path.append(str(script_dir.parent))
    else:
        # Add the current directory to the path
        sys.path.append(str(script_dir))
    
    # Try import again
    try:
        import utils
    except ImportError:
        print("ERROR: Could not import the utils module. Make sure utils.py is in the StratusScan directory.")
        sys.exit(1)

def check_dependencies():
    """
    Check if required dependencies are installed and offer to install them if missing.
    """
    required_packages = ['pandas', 'openpyxl', 'python-dateutil']
    missing_packages = []
    
    for package in required_packages:
        try:
            __import__(package)
            print(f"✓ {package} is already installed")
        except ImportError:
            missing_packages.append(package)
    
    if missing_packages:
        print(f"\nPackages required but not installed: {', '.join(missing_packages)}")
        response = input("Would you like to install these packages now? (y/n): ").lower()
        
        if response == 'y':
            import subprocess
            for package in missing_packages:
                print(f"Installing {package}...")
                try:
                    subprocess.check_call([sys.executable, "-m", "pip", "install", package])
                    print(f"✓ Successfully installed {package}")
                except Exception as e:
                    print(f"Error installing {package}: {e}")
                    print("Please install it manually with: pip install " + package)
                    return False
            return True
        else:
            print("Cannot proceed without required dependencies.")
            return False
    
    return True

# Now import the modules (after dependencies check in main())
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def print_title():
    """
    Print the script title banner and get account info.
    
    Returns:
        tuple: (account_id, account_name)
    """
    print("====================================================================")
    print("                  AWS RESOURCE SCANNER                              ")
    print("====================================================================")
    print("AWS ACCOUNT BILLING DATA EXPORT TOOL")
    print("====================================================================")
    print("Version: v1.1.0                                 Date: MAR-03-2025")
    print("====================================================================")
    
    # Get the current AWS account ID
    try:
        # Create a new STS client to get the current account ID
        sts_client = boto3.client('sts')
        # Get account ID from caller identity
        account_id = sts_client.get_caller_identity()['Account']
        # Map the account ID to an account name using utils module
        account_name = utils.get_account_name(account_id, default=account_id)
        
        print(f"Account ID: {account_id}")
        print(f"Account Name: {account_name}")
    except Exception as e:
        print(f"Could not determine account information: {e}")
        account_id = "UNKNOWN"
        account_name = "UNKNOWN-ACCOUNT"
    
    print("====================================================================")
    return account_id, account_name

def validate_date_input(date_input):
    """
    Validate user input for year or month-year.
    
    Args:
        date_input (str): User input string
        
    Returns:
        tuple: (is_valid, is_year_only, start_date, end_date)
    """
    # Define regex patterns for year and month-year formats
    year_pattern = r'^\d{4}$'  # YYYY
    month_year_pattern = r'^(0[1-9]|1[0-2])-\d{4}$'  # MM-YYYY
    
    if re.match(year_pattern, date_input):
        # Year format (YYYY)
        year = int(date_input)
        start_date = datetime.datetime(year, 1, 1)
        end_date = datetime.datetime(year, 12, 31)
        return True, True, start_date, end_date
    
    elif re.match(month_year_pattern, date_input):
        # Month-Year format (MM-YYYY)
        month, year = date_input.split('-')
        month = int(month)
        year = int(year)
        
        start_date = datetime.datetime(year, month, 1)
        # Calculate the last day of the month
        if month == 12:
            end_date = datetime.datetime(year + 1, 1, 1) - datetime.timedelta(days=1)
        else:
            end_date = datetime.datetime(year, month + 1, 1) - datetime.timedelta(days=1)
        
        return True, False, start_date, end_date
    
    else:
        return False, None, None, None

def validate_date_range(start_date, end_date):
    """
    Validate the date range against AWS Cost Explorer limitations.
    
    Args:
        start_date (datetime): Start date
        end_date (datetime): End date
        
    Returns:
        tuple: (is_valid, message)
    """
    today = datetime.datetime.now()
    
    # Cost Explorer data is available the next day
    latest_available_date = today - datetime.timedelta(days=1)
    
    # By default, Cost Explorer can go back 14 months (~426 days)
    earliest_available_date = today - datetime.timedelta(days=426)  
    
    # Check if end date is in the future
    if end_date > latest_available_date:
        end_date_str = end_date.strftime('%Y-%m-%d')
        latest_date_str = latest_available_date.strftime('%Y-%m-%d')
        return False, f"End date ({end_date_str}) is in the future. Latest available data is for {latest_date_str}."
    
    # Check if start date is too far in the past
    if start_date < earliest_available_date:
        start_date_str = start_date.strftime('%Y-%m-%d')
        earliest_date_str = earliest_available_date.strftime('%Y-%m-%d')
        return False, f"Start date ({start_date_str}) is too far in the past. AWS Cost Explorer only provides data for the last 14 months by default (from {earliest_date_str})."
    
    return True, "Date range is valid."

def get_billing_data(start_date, end_date):
    """
    Get billing data from AWS Cost Explorer API.
    
    Args:
        start_date (datetime): Start date
        end_date (datetime): End date
        
    Returns:
        dict: Billing data organized by month and service
    """
    # Convert dates to string format required by AWS API
    start_date_str = start_date.strftime('%Y-%m-%d')
    end_date_str = end_date.strftime('%Y-%m-%d')
    
    print(f"Fetching billing data from {start_date_str} to {end_date_str}...")
    
    # Create a Cost Explorer client
    ce_client = boto3.client('ce')
    
    try:
        # Use the cost explorer API to get cost and usage data
        response = ce_client.get_cost_and_usage(
            TimePeriod={
                'Start': start_date_str,
                'End': end_date_str
            },
            Granularity='MONTHLY',
            Metrics=['BlendedCost'],
            GroupBy=[
                {
                    'Type': 'DIMENSION',
                    'Key': 'SERVICE'
                }
            ]
        )
        
        # Organize the data by month and service
        billing_data = {}
        
        for result in response['ResultsByTime']:
            # Extract the month from the time period
            period_start = result['TimePeriod']['Start']
            month = datetime.datetime.strptime(period_start, '%Y-%m-%d').strftime('%Y-%m')
            
            # Initialize the month in the billing data if not already present
            if month not in billing_data:
                billing_data[month] = {}
            
            # Process each service and its cost
            for group in result['Groups']:
                service_name = group['Keys'][0]
                cost = float(group['Metrics']['BlendedCost']['Amount'])
                
                # Add the service cost to the month data
                billing_data[month][service_name] = cost
        
        return billing_data
        
    except ClientError as e:
        error_code = e.response.get('Error', {}).get('Code', '')
        error_message = e.response.get('Error', {}).get('Message', str(e))
        
        if error_code == 'ValidationException' and 'historical data' in error_message:
            print("\nError: AWS Cost Explorer cannot access historical data beyond the default 14 months.")
            print("This is a limitation of AWS Cost Explorer. To access older data:")
            print("1. You need to enable longer data retention in Cost Explorer settings")
            print("2. Or choose a more recent date range")
            sys.exit(1)
        else:
            print(f"\nError accessing Cost Explorer: {error_message}")
            sys.exit(1)

def create_excel_report(billing_data, account_name, date_suffix):
    """
    Create an Excel report with monthly billing data.
    
    Args:
        billing_data (dict): Billing data organized by month and service
        account_name (str): Name of AWS account for file naming
        date_suffix (str): Date suffix for filename
        
    Returns:
        str: Path to the created Excel file
    """
    # Create a new workbook
    wb = Workbook()
    
    # Remove the default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Define styles
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    # Sort months chronologically
    sorted_months = sorted(billing_data.keys())
    
    # Process each month
    for month in sorted_months:
        # Create a sheet for the month
        sheet_name = datetime.datetime.strptime(month, '%Y-%m').strftime('%b %Y')
        ws = wb.create_sheet(sheet_name)
        
        # Create headers
        ws['A1'] = 'Service'
        ws['B1'] = 'Cost (USD)'
        
        # Apply header styles
        ws['A1'].font = header_font
        ws['B1'].font = header_font
        ws['A1'].fill = header_fill
        ws['B1'].fill = header_fill
        
        # Get monthly data
        month_data = billing_data[month]
        
        # Sort services by cost (descending)
        sorted_services = sorted(month_data.items(), key=lambda x: x[1], reverse=True)
        
        # Add data rows
        row = 2
        total_cost = 0
        
        for service, cost in sorted_services:
            ws[f'A{row}'] = service
            ws[f'B{row}'] = cost
            ws[f'B{row}'].number_format = '$#,##0.00'
            total_cost += cost
            row += 1
        
        # Add total row
        row += 1
        ws[f'A{row}'] = 'Total'
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'] = total_cost
        ws[f'B{row}'].font = header_font
        ws[f'B{row}'].number_format = '$#,##0.00'
        
        # Adjust column widths
        for col in range(1, 3):
            column_letter = get_column_letter(col)
            max_length = 0
            for cell in ws[column_letter]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Generate filename using utils
    filename = utils.create_export_filename(
        account_name, 
        "billing", 
        date_suffix, 
        datetime.datetime.now().strftime("%m.%d.%Y")
    )
    
    # Get the full output path
    output_path = utils.get_output_filepath(filename)
    
    # Ensure the output directory exists
    output_dir = os.path.dirname(output_path)
    os.makedirs(output_dir, exist_ok=True)
    
    # Save the workbook
    wb.save(output_path)
    print(f"Excel report saved as: {output_path}")
    return output_path

def main():
    """
    Main function to run the script.
    """
    try:
        # Print title and get account info
        account_id, account_name = print_title()
        
        # Check dependencies
        if not check_dependencies():
            sys.exit(1)
            
        # Now import pandas after ensuring it's installed
        global pd
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        
        # Get user input for date range
        while True:
            date_input = input("\nIndicate a specific year (ex. \"2024\") or a specific month (ex. \"01-2025\"): ")
            is_valid, is_year_only, start_date, end_date = validate_date_input(date_input)
            
            if is_valid:
                # Validate date range against AWS limitations
                date_valid, message = validate_date_range(start_date, end_date)
                if date_valid:
                    break
                else:
                    print(f"Error: {message}")
                    
                    # Offer alternative suggestion
                    today = datetime.datetime.now()
                    suggestion_start = today - datetime.timedelta(days=365)  # Suggest last 12 months
                    suggestion_start = suggestion_start.replace(day=1)  # First day of that month
                    
                    if is_year_only:
                        # Suggest current year
                        suggestion = str(today.year)
                    else:
                        # Suggest format MM-YYYY for the last year
                        suggestion = suggestion_start.strftime("%m-%Y")
                    
                    print(f"Suggestion: Try a more recent date range, like \"{suggestion}\"")
            else:
                print("Invalid input format. Please try again.")
        
        # Get billing data
        billing_data = get_billing_data(start_date, end_date)
        
        # Determine output file name suffix
        if is_year_only:
            date_suffix = start_date.strftime('%Y')
        else:
            date_suffix = start_date.strftime('%m-%Y')
        
        # Create Excel report
        output_file = create_excel_report(billing_data, account_name, date_suffix)
        
        print("\nBilling data export completed successfully.")
        print(f"File saved to: {output_file}")
        
    except KeyboardInterrupt:
        print("\nOperation cancelled by user.")
        sys.exit(1)
    except Exception as e:
        print(f"\nAn error occurred: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
